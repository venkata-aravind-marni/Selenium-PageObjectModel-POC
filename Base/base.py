import pytest
import time
import unittest
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
import openpyxl
from dotenv import load_dotenv
load_dotenv()
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import sys
import os
sys.path.insert(0, os.getenv('locator_path'))
from locator import Locators

# from django.conf import settings
# from django.test import LiveServerTestCase
# from django.test.client import Client
# import pdb
# import var

# sys.path.insert(0, os.getenv('base'))
# # from base import Base
# from excel import Excel

# def create_session_store():
#     """ Creates a session storage object. """

#     from django.utils.importlib import import_module
#     engine = import_module(settings.SESSION_ENGINE)
#     # Implement a database session store object that will contain the session key.
#     store = engine.SessionStore()
#     store.save()
#     return store

class Base(unittest.TestCase):
    # arr=[]

    @classmethod
    def setUpClass(inst):       
        inst.driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
        inst.driver.maximize_window()
        inst.base_url = os.getenv('base_url')
        inst.driver.get(inst.base_url)
        time.sleep(1)

        inst.wb = openpyxl.load_workbook(os.getenv('excel_path'))
        inst.sheet = inst.wb["Login"]
        inst.sheet1 = inst.wb["Pages"]
        inst.sheet2 = inst.wb["User Guide"]
        inst.sheet3 = inst.wb["My Custom Ngage"]
        inst.sheet4 = inst.wb["Profile - Edit Profile"]
        inst.sheet5 = inst.wb["Research Library"]
        inst.sheet6 = inst.wb['Report Downloads']
        inst.sheet7 = inst.wb['Subscriptions']
        inst.sheet8 = inst.wb['Users List']
        inst.sheet9 = inst.wb['Users Add']
        inst.sheet10 = inst.wb['Users Edit']
        inst.sheet11 = inst.wb['Change Password']
        inst.sheet12 = inst.wb['News List']

    def demo_login(self):
        try:
            self.driver.find_element(By.ID, Locators.email).send_keys(self.sheet.cell(2, 1).value)
            self.driver.find_element(By.NAME, Locators.password).send_keys(self.sheet.cell(2, 2).value)
            self.driver.find_element(By.ID, Locators.login).click()
            time.sleep(1)
        except Exception as exception:
            self.driver.get_screenshot_as_file(os.getenv('screenshot_path'))
            function_name = "demo_login"
            self.final_mail(exception, function_name)                  

    def final_mail(self, err, function_name):
        fromaddr = "aravind@auroraelabs.com"
        subject = "error message"
        body = "Please find the error message and attachment"
        filename = "screenshot.png"
        # recipients = os.getenv('recipients')
        recipients = ['aravind@auroraelabs.com']
        scr_Path = os.getenv('screenshot_path')

        msg = MIMEMultipart()
        msg['From'] = fromaddr
        # msg['To'] = toaddr
        msg['To'] = ", ".join(recipients)
        msg['Subject'] = subject
        # html = """\
        # <html>
        #   <head></head>
        #   <body>
        #     <p>Please find the error message and attachment<br>
        #     <br>Error Message :  """ +str(err)+ """
        #     <br>Function Name :  """ +str(function_name)+ """
        #     </p>
        #   </body>
        # </html>
        # """
        html = """\<!DOCTYPE html><html><head><title>Marketngage Email</title><meta name='viewport' content='width=device-width, initial-scale=1.0'><style>.services{margin:0;padding:0;margin:0px;font-size:12px;font-weight:500;line-height:1.25;text-align:left;letter-spacing:0.5px;background-color:#055d9c;padding:5px;text-transform:uppercase;font-family:Roboto,RobotoDraft,Helvetica,Arial,sans-serif;color:white}.services-1{margin:0;padding:0;margin:0px;font-size:11px;font-weight:500;line-height:1.25;text-align:left;letter-spacing:0.5px;background-color:#055d9c;padding:5px;text-transform:uppercase;font-family:Roboto,RobotoDraft,Helvetica,Arial,sans-serif;color:white}.for-mobile{display:none}@media only screen and (max-width: 576px){.for-pc{display:none}}@media only screen and (max-width: 576px){.for-mobile{display:block}}</style></head><body style='width: 100%;height: auto;margin: 0;padding: 0;background-color: #ffffff;'><div class='container' style='width: 100%;margin-right: auto;margin-left: auto;background-color: #eff2f6;'> <a href='member.marketngage.com' target='_blank'><div class='header' style='background-color: #055d9c;text-align: center;'> <img src='member.marketngage.com/images/favicon.png' alt='' style='padding: 20px;'></div> </a><div style='padding: 20px;'><table width='100%' cellspacing='0' cellpadding='0' style='padding: 10px;text-align: center;background-color: #fff;padding:10% 0%;'><tbody><tr style='margin:0;padding:0'><td style='margin:0;padding:0;text-align:center'><p style='font-size: 15px;padding: 15px 10px;font-weight: 400;line-height: 1.55;text-align: left;letter-spacing: 0.5px;color: #055d9c;font-family: Roboto,RobotoDraft,Helvetica,Arial,sans-serif;'>Please find the error message and attachment<br><br>Error Message :  """ +str(err)+ """<br>Function Name :  """ +str(function_name)+ """</p></td></tr></tbody></table></div><div style='margin:0;padding:0 15px;background:#eff2f6;'><table width='100%' cellspacing='0' cellpadding='0' style='padding: 10px;text-align: center;background-color: #055d9c;'><tbody><tr style='margin:0;padding:0'><td style='margin:0;padding:0;text-align:center'><p style='font-size: 10px;padding: 8px;font-weight: 500;line-height: 1.55;text-align: center;letter-spacing: 0.5px;color: #fff;font-family: Roboto,RobotoDraft,Helvetica,Arial,sans-serif; margin: 0px;'>© Copyright 2021, Marketngage, Inc. All Rights Reserved.</p></td></tr></tbody></table></div></div></body></html>"""

        msg.attach(MIMEText(html, 'html'))
        attachment = open(scr_Path, "rb")

        p = MIMEBase('application', 'octet-stream')
        p.set_payload((attachment).read())
        encoders.encode_base64(p)
        p.add_header('Content-Disposition', "attachment; filename= %s" % filename)
        msg.attach(p)

        s = smtplib.SMTP('smtp.gmail.com', 587)
        s.starttls()
        s.login(fromaddr, "tpmmlcsquprvwdyo")
        # text = msg.as_string()
        try:
            s.sendmail(fromaddr, recipients, msg.as_string())
            print("Error details sent email to", recipients)
        except Exception as e:
            print("Error: unable to send email")
            # print(e)
        s.quit()    

    @classmethod
    def tearDownClass(inst):
        # close the browser window
        inst.driver.quit()
    

    # def test_welcome_page(self):
    #     #pdb.set_trace()
    #     # Create a session storage object.
    #     session_store = create_session_store()
    #     # In pdb, you can do 'session_store.session_key' to view the session key just created.

    #     # Create a session object from the session store object.
    #     session_items = session_store

    #     # Add a session key/value pair.
    #     session_items['arr'] = []
    #     session_items.save()

    #     # Go to the correct domain.
    #     self.browser.get(self.live_server_url)

    #     # Add the session key to the cookie that will be sent back to the server.
    #     # self.browser.add_cookie({'name': settings.SESSION_COOKIE_NAME, 'value': session_store.session_key})
    #     # In pdb, do 'self.browser.get_cookies() to verify that it's there.'

    #     # The client sends a request to the view that's expecting the session item.
    #     # self.browser.get(self.live_server_url + '/signup/')
    #     # body = self.browser.find_element_by_tag_name('body')
    #     # self.assertIn('Welcome', body.text)    