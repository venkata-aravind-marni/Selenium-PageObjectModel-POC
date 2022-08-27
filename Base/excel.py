import unittest
import os
import sys
from dotenv import load_dotenv
load_dotenv()
sys.path.insert(0, os.getenv('basepy_path'))
from base import Base
from selenium.webdriver.common.by import By
import time
from selenium.webdriver.support.ui import Select
import openpyxl

sys.path.insert(0, os.getenv('locator_path'))
from locator import Locators

class Excel(unittest.TestCase):
  
    def test_excel(inst):
        # try:
        print("running:test_excel")

        inst.wb = openpyxl.load_workbook(os.getenv('excel_path'))

        inst.sheet = inst.wb["Login"]
        inst.sheet.delete_cols(4)
        inst.sheet.cell(1, 4).value = "Result"
        for row in range(2,inst.sheet.max_row + 1):
            inst.sheet.cell(row, 4).value = "-"
        inst.wb.save("member.xlsx")

        inst.sheet1 = inst.wb["Pages"]
        inst.sheet1.delete_cols(2)
        inst.sheet1.cell(1, 2).value = "Result"
        for row in range(2,inst.sheet1.max_row + 1):
            inst.sheet1.cell(row, 2).value = "-"
        inst.wb.save("member.xlsx")

        inst.sheet2 = inst.wb["User Guide"]
        inst.sheet2.delete_cols(4)
        inst.sheet2.cell(1, 4).value = "Result"
        for row in range(2,inst.sheet2.max_row + 1):
            inst.sheet2.cell(row, 4).value = "-"
        inst.wb.save("member.xlsx")

        inst.sheet3 = inst.wb["My Custom Ngage"]
        inst.sheet3.delete_cols(5)
        inst.sheet3.cell(1, 5).value = "Result"
        for row in range(2,inst.sheet3.max_row + 1):
            inst.sheet3.cell(row, 5).value = "-"
        inst.wb.save("member.xlsx")

        inst.sheet4 = inst.wb["Profile - Edit Profile"]
        inst.sheet4.delete_cols(8)
        inst.sheet4.cell(1, 8).value = "Result"
        for row in range(2,inst.sheet4.max_row + 1):
            inst.sheet4.cell(row, 8).value = "-"
        inst.wb.save("member.xlsx")

        inst.sheet5 = inst.wb["Research Library"]
        inst.sheet5.delete_cols(5)
        inst.sheet5.cell(1,5).value = "Result"
        for row in range(2,inst.sheet5.max_row + 1):
            inst.sheet5.cell(row, 5).value = "-"
        inst.wb.save("member.xlsx")

        inst.sheet6 = inst.wb['Report Downloads']
        inst.sheet6.delete_cols(2)
        inst.sheet6.cell(1,2).value = "Result"
        for row in range(2,inst.sheet6.max_row + 1):
            inst.sheet6.cell(row, 2).value = "-"
        inst.wb.save("member.xlsx")

        inst.sheet7 = inst.wb['Subscriptions']
        inst.sheet7.delete_cols(2)
        inst.sheet7.cell(1,2).value = "Result"
        for row in range(2,inst.sheet7.max_row + 1):
            inst.sheet7.cell(row, 2).value = "-"
        inst.wb.save("member.xlsx")

        inst.sheet8 = inst.wb['Users List']
        inst.sheet8.delete_cols(4)
        inst.sheet8.cell(1,4).value = "Result"
        for row in range(2,inst.sheet8.max_row + 1):
            inst.sheet8.cell(row, 4).value = "-"
        inst.wb.save("member.xlsx")

        inst.sheet9 = inst.wb['Users Add']
        inst.sheet9.delete_cols(12)
        inst.sheet9.cell(1,12).value = "Result"
        for row in range(2,inst.sheet9.max_row + 1):
            inst.sheet9.cell(row, 12).value = "-"
        inst.wb.save("member.xlsx")

        inst.sheet10 = inst.wb['Users Edit']
        inst.sheet10.delete_cols(12)
        inst.sheet10.cell(1,12).value = "Result"
        for row in range(2,inst.sheet10.max_row + 1):
            inst.sheet10.cell(row, 12).value = "-"
        inst.wb.save("member.xlsx")

        inst.sheet11 = inst.wb["Change Password"]
        inst.sheet11.delete_cols(6)
        inst.sheet11.cell(1, 6).value = "Result"
        for row in range(2,inst.sheet11.max_row + 1):
            inst.sheet11.cell(row, 6).value = "-"
        inst.wb.save("member.xlsx")

        inst.sheet12 = inst.wb["News List"]
        inst.sheet12.delete_cols(4)
        inst.sheet12.cell(1, 4).value = "Result"
        for row in range(2,inst.sheet12.max_row + 1):
            inst.sheet12.cell(row, 4).value = "-"
        inst.wb.save("member.xlsx")



        # except Exception as exception:
        #     self.driver.get_screenshot_as_file(os.getenv('screenshot_path'))
        #     function_name = "test_ReportDownloads"
        #     self.final_mail(exception, function_name) 

         